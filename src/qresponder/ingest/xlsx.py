"""XLSX loader (§6, Stage A).

Emits every non-empty cell as `Sheet!Coord = value` plus style signals
(fill color, bold, merged-range membership) because color and merges carry
meaning. Merged ranges are walked via `ws.merged_cells`; the value lives on the
top-left cell, and other members are flagged as merged so the extractor knows
they are visually part of the same logical cell.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .ir import Document, Element

# openpyxl reports "no fill" with these fgColor rgb values; treat as no signal.
_DEFAULT_FILLS = {None, "00000000", "FFFFFFFF"}


def _fill_color(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    # Theme colors come back as ints/objects; only surface explicit rgb strings.
    if isinstance(rgb, str) and rgb not in _DEFAULT_FILLS:
        return rgb
    return None


def load_xlsx(path: str | Path) -> Document:
    p = Path(path)
    # data_only=True -> resolved values rather than formula strings.
    wb = openpyxl.load_workbook(p, data_only=True)
    elements: list[Element] = []

    for ws in wb.worksheets:
        # Map every cell coordinate inside a merged range, and find the anchors.
        merged_members: set[str] = set()
        merged_anchors: set[str] = set()
        for mrange in ws.merged_cells.ranges:
            anchor = f"{get_column_letter(mrange.min_col)}{mrange.min_row}"
            merged_anchors.add(anchor)
            for row in range(mrange.min_row, mrange.max_row + 1):
                for col in range(mrange.min_col, mrange.max_col + 1):
                    merged_members.add(f"{get_column_letter(col)}{row}")

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None or str(value).strip() == "":
                    continue
                coord = cell.coordinate
                style: dict = {}
                font = cell.font
                if font is not None and font.bold:
                    style["bold"] = True
                fill = _fill_color(cell)
                if fill:
                    style["fill"] = fill
                if coord in merged_members:
                    style["merged"] = True
                    if coord in merged_anchors:
                        style["merge_anchor"] = True

                elements.append(
                    Element(
                        kind="cell",
                        location=f"{ws.title}!{coord}",
                        text=str(value),
                        style=style,
                    )
                )

    wb.close()
    return Document(source_file=str(p), file_type="xlsx", elements=elements)
