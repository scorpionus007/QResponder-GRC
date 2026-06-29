"""Phase 0 output (§15): a clean answered.xlsx + results.json.

Columns: Question / Answer / Citation / Confidence / Status / Reason. This is a
fresh, readable artifact — NOT the format-perfect write-back into the original
file (that's Phase 2, see writeback.py).
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from ..models import QuestionnaireResult, Status

_HEADERS = ["#", "Question", "Answer", "Citations", "Confidence", "Status", "Reason", "Tier"]
_REVIEW_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

# Visible placeholder marker for NEEDS_REVIEW items (Phase 7 C).
DEFAULT_REVIEW_MARKER = "⚠ NEEDS REVIEW: {reason}"


def review_marker(result, template: str = DEFAULT_REVIEW_MARKER) -> str:
    reason = (result.review_reason.value or "review").replace("_", " ")
    try:
        return template.format(reason=reason)
    except (KeyError, IndexError, ValueError):
        return f"⚠ NEEDS REVIEW: {reason}"


def _citation_text(result) -> str:
    parts = []
    for c in result.citations:
        snippet = c.snippet if len(c.snippet) <= 200 else c.snippet[:197] + "..."
        parts.append(f"[{c.source}] {snippet}")
    return "\n".join(parts)


def write_json(result: QuestionnaireResult, path: str | Path) -> Path:
    p = Path(path)
    p.write_text(result.model_dump_json(indent=2), encoding="utf-8")
    return p


def write_xlsx(result: QuestionnaireResult, path: str | Path, review_markers: bool = True) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"

    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for i, r in enumerate(result.results, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.question_text)
        if r.status == Status.NEEDS_REVIEW and review_markers:
            extra = f" — {r.missing_info}" if r.missing_info else ""
            answer_cell = review_marker(r) + extra
        else:
            answer_cell = r.answer or (r.missing_info or "")
        ws.cell(row=row, column=3, value=answer_cell)
        ws.cell(row=row, column=4, value=_citation_text(r))
        ws.cell(row=row, column=5, value=r.confidence.value)
        ws.cell(row=row, column=6, value=r.status.value)
        ws.cell(row=row, column=7, value=r.review_reason.value)
        ws.cell(row=row, column=8, value=r.source_tier)
        if r.status == Status.NEEDS_REVIEW:
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=row, column=col).fill = _REVIEW_FILL

    # Reasonable column widths for readability.
    widths = {1: 4, 2: 50, 3: 60, 4: 50, 5: 11, 6: 14, 7: 16, 8: 6}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    p = Path(path)
    wb.save(p)
    return p


def write_all(result: QuestionnaireResult, out_dir: str | Path, review_markers: bool = True) -> dict[str, Path]:
    """Write answered.xlsx, results.json, and review.md to out_dir."""
    from .review import write_review_md

    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    return {
        "xlsx": write_xlsx(result, d / "answered.xlsx", review_markers=review_markers),
        "json": write_json(result, d / "results.json"),
        "review": write_review_md(result, d / "review.md"),
    }
