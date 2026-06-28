"""End-to-end pipeline + output smoke test (mock provider, no network).

This exercises the full Phase 0 acceptance path (§17): ingest -> extract ->
orchestrate -> write answered.xlsx + results.json + review.md, asserting the
guardrails hold across the whole flow.
"""

import json
from pathlib import Path

import openpyxl

from qresponder.config import Config
from qresponder.core.pipeline import run_pipeline
from qresponder.llm.mock import MockProvider
from qresponder.models import Status
from qresponder.output.writer import write_all

FIX = Path(__file__).parent / "fixtures"


def test_end_to_end_with_outputs(tmp_path):
    cfg = Config(llm_provider="mock", kb_mode="in_context")
    result = run_pipeline(
        questionnaire_path=str(FIX / "sample.xlsx"),
        kb_dir=str(FIX / "kb"),
        qa_path=str(FIX / "qa.yaml"),
        config=cfg,
        scope_tags=["soc2"],
        provider=MockProvider(),
    )

    assert result.results, "pipeline produced no results"

    # GUARDRAIL: no ANSWERED without a citation; no fabrication.
    for r in result.results:
        if r.status == Status.ANSWERED:
            assert r.citations, f"{r.question_id} answered without citation"

    # Tier-1 reuse happened for at least one question.
    assert any(r.source_tier == 1 for r in result.results)

    # The unsupported "backup retention" question is flagged, not fabricated.
    retention = next((r for r in result.results if "retention" in r.question_text.lower()), None)
    assert retention is not None
    assert retention.status == Status.NEEDS_REVIEW

    # Write outputs and verify all three files exist and are well-formed.
    paths = write_all(result, tmp_path)
    assert paths["xlsx"].exists()
    assert paths["json"].exists()
    assert paths["review"].exists()

    data = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert data["source_file"].endswith("sample.xlsx")
    assert len(data["results"]) == len(result.results)

    wb = openpyxl.load_workbook(paths["xlsx"])
    ws = wb.active
    assert ws.cell(row=1, column=2).value == "Question"
    assert ws.max_row == len(result.results) + 1

    review = paths["review"].read_text(encoding="utf-8")
    assert "Review report" in review
    assert "draft" in review.lower()
